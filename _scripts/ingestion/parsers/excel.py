"""
Excel file parser (.xlsx, .xls).

Extracts data from spreadsheets for analysis.

Copyright (c) 2025 Brent Lefebure / EhkoLabs
Licensed under AGPLv3
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from .base import BaseParser
from ..types import ParsedContent


class ExcelParser(BaseParser):
    """
    Parse Excel spreadsheets (.xlsx, .xls).
    
    Extracts:
    - Cell data as text
    - Sheet names
    - Structured table data
    """
    
    def get_extensions(self) -> List[str]:
        return [".xlsx", ".xls", ".xlsm"]
    
    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.get_extensions()
    
    def get_file_type(self) -> str:
        return "excel"
    
    def parse(self, path: Path) -> ParsedContent:
        """Parse Excel file and extract text content."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ParsedContent(
                text="[Excel parsing requires openpyxl: pip install openpyxl]",
                title=path.stem,
                metadata={"error": "openpyxl not installed"}
            )
        
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            return ParsedContent(
                text=f"[Error reading Excel file: {e}]",
                title=path.stem,
                metadata={"error": str(e)}
            )
        
        all_text = []
        sheet_data = []
        total_rows = 0
        total_cells = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            all_text.append(f"\n{'='*60}")
            all_text.append(f"Sheet: {sheet_name}")
            all_text.append("="*60 + "\n")
            
            sheet_rows = []
            row_count = 0
            
            for row in sheet.iter_rows():
                row_values = []
                has_content = False
                
                for cell in row:
                    value = cell.value
                    if value is not None:
                        has_content = True
                        # Convert datetime to string
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        row_values.append(str(value))
                        total_cells += 1
                    else:
                        row_values.append("")
                
                if has_content:
                    sheet_rows.append(row_values)
                    row_count += 1
                    # Format as tab-separated for text output
                    all_text.append("\t".join(row_values))
            
            total_rows += row_count
            
            sheet_data.append({
                "name": sheet_name,
                "rows": row_count,
                "data": sheet_rows[:100],  # Limit stored data
            })
        
        wb.close()
        
        metadata = {
            "format": "excel",
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "total_rows": total_rows,
            "total_cells": total_cells,
            "sheets": sheet_data,
        }
        
        return ParsedContent(
            text="\n".join(all_text),
            title=f"{path.stem} ({len(wb.sheetnames)} sheets)",
            metadata=metadata,
        )


__all__ = ["ExcelParser"]
