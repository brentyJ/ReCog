"""
ReCog - File Validation v1.0

Copyright (c) 2025 Brent Lefebure / EhkoLabs
Licensed under AGPLv3 - See LICENSE in repository root

Validates uploaded files before processing:
- File size limits
- Magic byte detection (actual file type)
- Empty/corrupted file detection
- Text extractability check
"""

import logging
from pathlib import Path
from typing import Optional, Dict, Any, Tuple, BinaryIO
from dataclasses import dataclass

from .errors import (
    FileTooLargeError,
    EmptyFileError,
    CorruptedFileError,
    UnsupportedFileTypeError,
    NoExtractableTextError,
)

logger = logging.getLogger(__name__)


# =============================================================================
# CONFIGURATION
# =============================================================================

# Default file size limit (10MB as per task spec)
DEFAULT_MAX_SIZE_MB = 10

# Minimum file size to consider non-empty (bytes)
MIN_FILE_SIZE = 10

# Minimum extractable text length to consider valid
MIN_TEXT_LENGTH = 50


# =============================================================================
# MAGIC BYTE SIGNATURES
# =============================================================================
# Maps file types to their magic byte signatures

MAGIC_SIGNATURES = {
    # PDF - starts with %PDF
    "application/pdf": [
        (b"%PDF", 0),
    ],
    # Microsoft Office (OOXML) - ZIP-based formats
    "application/vnd.openxmlformats-officedocument": [
        (b"PK\x03\x04", 0),  # ZIP signature
    ],
    # Plain text files have no signature, detected by absence of binary
    "text/plain": [],
    # JSON - typically starts with { or [
    "application/json": [
        (b"{", 0),
        (b"[", 0),
        (b" {", 0),  # Whitespace prefix
        (b" [", 0),
        (b"\n{", 0),
        (b"\n[", 0),
    ],
    # XML/HTML
    "text/xml": [
        (b"<?xml", 0),
        (b"<!", 0),  # DOCTYPE
    ],
    # Email formats
    "message/rfc822": [  # .eml files
        (b"From:", 0),
        (b"Return-Path:", 0),
        (b"Received:", 0),
        (b"MIME-Version:", 0),
        (b"Date:", 0),
    ],
    # CSV - no magic, but shouldn't have binary
    "text/csv": [],
    # Markdown - no magic, treated as text
    "text/markdown": [],
}

# Map extensions to expected MIME types
EXTENSION_TO_MIME = {
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".json": "application/json",
    ".pdf": "application/pdf",
    ".csv": "text/csv",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".eml": "message/rfc822",
    ".msg": "application/vnd.ms-outlook",
    ".xml": "text/xml",
    ".mbox": "application/mbox",
}

# Supported MIME types for processing
SUPPORTED_MIME_TYPES = {
    "text/plain",
    "text/markdown",
    "text/csv",
    "text/xml",
    "application/json",
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "message/rfc822",
    "application/mbox",
}


@dataclass
class ValidationResult:
    """Result of file validation."""
    valid: bool
    mime_type: Optional[str] = None
    extension: Optional[str] = None
    size_bytes: int = 0
    size_mb: float = 0.0
    error: Optional[str] = None
    warning: Optional[str] = None


class FileValidator:
    """
    Validates files before processing.

    Checks:
    1. File size within limits
    2. File type via magic bytes
    3. File not empty or corrupted
    4. Text is extractable

    Usage:
        validator = FileValidator(max_size_mb=10)
        result = validator.validate_file(file_path)
        if not result.valid:
            raise result.error
    """

    def __init__(self, max_size_mb: float = DEFAULT_MAX_SIZE_MB):
        """
        Initialize validator.

        Args:
            max_size_mb: Maximum allowed file size in MB
        """
        self.max_size_mb = max_size_mb
        self.max_size_bytes = int(max_size_mb * 1024 * 1024)

    def validate_file(self, file_path: Path | str) -> ValidationResult:
        """
        Validate a file for processing.

        Args:
            file_path: Path to the file

        Returns:
            ValidationResult with valid=True or error details

        Raises:
            FileTooLargeError: File exceeds size limit
            EmptyFileError: File is empty
            CorruptedFileError: File appears corrupted
            UnsupportedFileTypeError: File type not supported
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise CorruptedFileError("File does not exist")

        # Get file info
        size_bytes = file_path.stat().st_size
        size_mb = size_bytes / (1024 * 1024)
        extension = file_path.suffix.lower()

        # Check 1: File size
        if size_bytes > self.max_size_bytes:
            raise FileTooLargeError(size_mb, self.max_size_mb)

        # Check 2: Empty file
        if size_bytes < MIN_FILE_SIZE:
            raise EmptyFileError()

        # Check 3: Detect actual file type via magic bytes
        try:
            mime_type = self._detect_mime_type(file_path, extension)
        except Exception as e:
            raise CorruptedFileError(f"Could not read file: {e}")

        # Check 4: Supported type
        if mime_type not in SUPPORTED_MIME_TYPES:
            raise UnsupportedFileTypeError(
                mime_type or "unknown",
                list(SUPPORTED_MIME_TYPES)
            )

        # Check 5: Validate content integrity based on type
        self._validate_content(file_path, mime_type)

        return ValidationResult(
            valid=True,
            mime_type=mime_type,
            extension=extension,
            size_bytes=size_bytes,
            size_mb=size_mb,
        )

    def validate_upload(
        self,
        file_storage,
        filename: str,
    ) -> ValidationResult:
        """
        Validate an uploaded file (from request.files).

        Args:
            file_storage: Flask FileStorage object
            filename: Original filename

        Returns:
            ValidationResult

        Raises:
            FileTooLargeError, EmptyFileError, etc.
        """
        # Get size by seeking to end
        file_storage.seek(0, 2)
        size_bytes = file_storage.tell()
        file_storage.seek(0)

        size_mb = size_bytes / (1024 * 1024)
        extension = Path(filename).suffix.lower()

        # Check size
        if size_bytes > self.max_size_bytes:
            raise FileTooLargeError(size_mb, self.max_size_mb)

        # Check empty
        if size_bytes < MIN_FILE_SIZE:
            raise EmptyFileError()

        # Detect type from content
        mime_type = self._detect_mime_type_from_stream(file_storage, extension)

        # Check supported
        if mime_type not in SUPPORTED_MIME_TYPES:
            raise UnsupportedFileTypeError(
                mime_type or extension or "unknown",
                list(SUPPORTED_MIME_TYPES)
            )

        return ValidationResult(
            valid=True,
            mime_type=mime_type,
            extension=extension,
            size_bytes=size_bytes,
            size_mb=size_mb,
        )

    def _detect_mime_type(self, file_path: Path, extension: str) -> str:
        """Detect MIME type from file content using magic bytes."""
        with open(file_path, "rb") as f:
            return self._detect_mime_type_from_stream(f, extension)

    def _detect_mime_type_from_stream(self, stream: BinaryIO, extension: str) -> str:
        """Detect MIME type from a file stream."""
        # Read first 1KB for signature detection
        stream.seek(0)
        header = stream.read(1024)
        stream.seek(0)

        if len(header) == 0:
            raise EmptyFileError()

        # Check for binary content (indicates non-text file)
        is_binary = self._is_binary(header)

        # Check magic signatures
        for mime_type, signatures in MAGIC_SIGNATURES.items():
            for signature, offset in signatures:
                if header[offset:offset + len(signature)] == signature:
                    # Special handling for ZIP-based formats (OOXML)
                    if mime_type == "application/vnd.openxmlformats-officedocument":
                        # Determine specific type from extension
                        if extension == ".xlsx":
                            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        elif extension == ".docx":
                            return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                        else:
                            return mime_type
                    return mime_type

        # Fall back to extension mapping for text-like files
        if not is_binary:
            expected_mime = EXTENSION_TO_MIME.get(extension)
            if expected_mime:
                return expected_mime
            # Default to text/plain for non-binary
            return "text/plain"

        # Binary file with no recognized signature
        expected_mime = EXTENSION_TO_MIME.get(extension)
        if expected_mime:
            return expected_mime

        return "application/octet-stream"

    def _is_binary(self, data: bytes) -> bool:
        """Check if data contains binary content."""
        # Null bytes indicate binary
        if b"\x00" in data:
            return True

        # Check for high proportion of non-text bytes
        text_chars = set(bytes(range(32, 127))) | {9, 10, 13}  # Printable + tab, newline, CR
        non_text = sum(1 for b in data if b not in text_chars)

        # If more than 30% non-text, consider binary
        return (non_text / len(data)) > 0.3 if data else False

    def _validate_content(self, file_path: Path, mime_type: str) -> None:
        """
        Validate file content based on type.

        Raises appropriate errors for corrupted/invalid content.
        """
        if mime_type == "application/pdf":
            self._validate_pdf(file_path)
        elif mime_type in ("text/plain", "text/markdown", "text/csv"):
            self._validate_text(file_path)
        elif mime_type == "application/json":
            self._validate_json(file_path)
        elif "spreadsheet" in mime_type or mime_type == "application/vnd.ms-excel":
            self._validate_excel(file_path)

    def _validate_pdf(self, file_path: Path) -> None:
        """Validate PDF file can be read and has text."""
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(file_path))

            if len(reader.pages) == 0:
                raise CorruptedFileError("PDF has no pages")

            # Try to extract text from first few pages
            text = ""
            for i, page in enumerate(reader.pages[:3]):
                try:
                    text += page.extract_text() or ""
                except Exception:
                    pass

            # Check if we got any meaningful text
            text = text.strip()
            if len(text) < MIN_TEXT_LENGTH:
                raise NoExtractableTextError("PDF")

        except NoExtractableTextError:
            raise
        except CorruptedFileError:
            raise
        except Exception as e:
            raise CorruptedFileError(f"PDF is corrupted or encrypted: {str(e)[:100]}")

    def _validate_text(self, file_path: Path) -> None:
        """Validate text file has readable content."""
        try:
            # Try common encodings
            content = None
            for encoding in ["utf-8", "utf-16", "latin-1"]:
                try:
                    content = file_path.read_text(encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if content is None:
                raise CorruptedFileError("Could not decode text file with any common encoding")

            if len(content.strip()) < MIN_TEXT_LENGTH:
                raise NoExtractableTextError("text file")

        except NoExtractableTextError:
            raise
        except CorruptedFileError:
            raise
        except Exception as e:
            raise CorruptedFileError(f"Could not read text file: {e}")

    def _validate_json(self, file_path: Path) -> None:
        """Validate JSON file is parseable."""
        import json
        try:
            content = file_path.read_text(encoding="utf-8")
            json.loads(content)

            if len(content.strip()) < MIN_TEXT_LENGTH:
                raise NoExtractableTextError("JSON file")

        except json.JSONDecodeError as e:
            raise CorruptedFileError(f"Invalid JSON: {str(e)[:100]}")
        except NoExtractableTextError:
            raise
        except Exception as e:
            raise CorruptedFileError(f"Could not read JSON file: {e}")

    def _validate_excel(self, file_path: Path) -> None:
        """Validate Excel file can be read."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), read_only=True, data_only=True)

            # Check we have at least one sheet with data
            has_data = False
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(max_row=10, values_only=True):
                    if any(cell is not None for cell in row):
                        has_data = True
                        break
                if has_data:
                    break

            wb.close()

            if not has_data:
                raise NoExtractableTextError("Excel file")

        except NoExtractableTextError:
            raise
        except Exception as e:
            raise CorruptedFileError(f"Could not read Excel file: {str(e)[:100]}")


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

_default_validator: Optional[FileValidator] = None


def get_validator(max_size_mb: float = DEFAULT_MAX_SIZE_MB) -> FileValidator:
    """Get or create a file validator instance."""
    global _default_validator
    if _default_validator is None:
        _default_validator = FileValidator(max_size_mb)
    return _default_validator


def validate_file(file_path: Path | str, max_size_mb: float = DEFAULT_MAX_SIZE_MB) -> ValidationResult:
    """
    Validate a file for processing.

    Args:
        file_path: Path to file
        max_size_mb: Maximum file size in MB

    Returns:
        ValidationResult

    Raises:
        FileTooLargeError, EmptyFileError, CorruptedFileError,
        UnsupportedFileTypeError, NoExtractableTextError
    """
    validator = FileValidator(max_size_mb)
    return validator.validate_file(file_path)


def validate_upload(file_storage, filename: str, max_size_mb: float = DEFAULT_MAX_SIZE_MB) -> ValidationResult:
    """
    Validate an uploaded file.

    Args:
        file_storage: Flask FileStorage object
        filename: Original filename
        max_size_mb: Maximum file size in MB

    Returns:
        ValidationResult
    """
    validator = FileValidator(max_size_mb)
    return validator.validate_upload(file_storage, filename)


# =============================================================================
# MODULE EXPORTS
# =============================================================================

__all__ = [
    "FileValidator",
    "ValidationResult",
    "validate_file",
    "validate_upload",
    "get_validator",
    "DEFAULT_MAX_SIZE_MB",
    "SUPPORTED_MIME_TYPES",
]
